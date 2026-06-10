import os
from PySide6.QtWidgets import *
from database import conectar

from reportlab.pdfgen import canvas
import pandas as pd


# ==========================
# FUNCIÓN PARA OBTENER ESCRITORIO
# ==========================
def obtener_escritorio():
    """Obtiene la ruta del escritorio"""
    return os.path.join(os.path.expanduser("~"), "Desktop")


class ReportesWidget(QWidget):

    def __init__(self):
        super().__init__()

        self.setStyleSheet("""
            QWidget{
                background:#0F172A;
                color:white;
                font-family:'Segoe UI';
            }

            QLabel#titulo{
                font-size:30px;
                font-weight:700;
                color:white;
            }

            QLabel#subtitulo{
                color:#CBD5E1;
                font-size:14px;
            }

            QFrame{
                background:#111827;
                border:1px solid #1E293B;
                border-radius:18px;
            }

            QPushButton{
                border:none;
                border-radius:12px;
                padding:14px;
                color:white;
                font-size:14px;
                font-weight:bold;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)

        # ==========================
        # HEADER
        # ==========================
        header = QFrame()
        header.setFixedHeight(130)

        header.setStyleSheet("""
            QFrame{
                background:qlineargradient(
                    x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1E3A5F,
                    stop:0.5 #172554,
                    stop:1 #1E3A8A
                );
                border-radius:20px;
            }
        """)

        headerLayout = QVBoxLayout()

        titulo = QLabel(" Centro de Reportes")
        titulo.setObjectName("titulo")

        subtitulo = QLabel(
            "Genera reportes PDF y Excel de los alumnos registrados"
        )
        subtitulo.setObjectName("subtitulo")

        headerLayout.addStretch()
        headerLayout.addWidget(titulo)
        headerLayout.addWidget(subtitulo)
        headerLayout.addStretch()

        header.setLayout(headerLayout)

        layout.addWidget(header)

        # ==========================
        # TARJETAS
        # ==========================
        cards = QHBoxLayout()
        cards.setSpacing(20)

        # --------------------------
        # PDF ALUMNOS
        # --------------------------
        pdfCard = QFrame()

        pdfLayout = QVBoxLayout()

        pdfTitulo = QLabel(" Reporte de Alumnos")
        pdfTitulo.setStyleSheet("font-size:22px;font-weight:bold;color:white;")

        pdfDesc = QLabel("Genera un documento PDF con todos los alumnos registrados.")
        pdfDesc.setWordWrap(True)
        pdfDesc.setStyleSheet("color:#94A3B8;font-size:13px;")

        self.btnPDF = QPushButton(" Generar PDF Alumnos")
        self.btnPDF.setStyleSheet("""
            QPushButton{background:#DC2626;border:none;border-radius:12px;padding:14px;font-weight:bold;}
            QPushButton:hover{background:#EF4444;}
        """)

        pdfLayout.addWidget(pdfTitulo)
        pdfLayout.addWidget(pdfDesc)
        pdfLayout.addStretch()
        pdfLayout.addWidget(self.btnPDF)

        pdfCard.setLayout(pdfLayout)

        # --------------------------
        # PDF MATERIAS
        # --------------------------
        materiasCard = QFrame()

        materiasLayout = QVBoxLayout()

        materiasTitulo = QLabel(" Reporte de Materias")
        materiasTitulo.setStyleSheet("font-size:22px;font-weight:bold;color:white;")

        materiasDesc = QLabel("Genera un PDF con las materias y alumnos inscritos.")
        materiasDesc.setWordWrap(True)
        materiasDesc.setStyleSheet("color:#94A3B8;font-size:13px;")

        self.btnPDFMaterias = QPushButton(" Generar PDF Materias")
        self.btnPDFMaterias.setStyleSheet("""
            QPushButton{background:#7C3AED;border:none;border-radius:12px;padding:14px;font-weight:bold;}
            QPushButton:hover{background:#8B5CF6;}
        """)

        materiasLayout.addWidget(materiasTitulo)
        materiasLayout.addWidget(materiasDesc)
        materiasLayout.addStretch()
        materiasLayout.addWidget(self.btnPDFMaterias)

        materiasCard.setLayout(materiasLayout)

        # --------------------------
        # EXCEL
        # --------------------------
        excelCard = QFrame()

        excelLayout = QVBoxLayout()

        excelTitulo = QLabel(" Reporte Excel")
        excelTitulo.setStyleSheet("font-size:22px;font-weight:bold;color:white;")

        excelDesc = QLabel("Exporta todos los alumnos a Excel.")
        excelDesc.setWordWrap(True)
        excelDesc.setStyleSheet("color:#94A3B8;font-size:13px;")

        self.btnExcel = QPushButton(" Exportar Excel")
        self.btnExcel.setStyleSheet("""
            QPushButton{background:#16A34A;border:none;border-radius:12px;padding:14px;font-weight:bold;}
            QPushButton:hover{background:#22C55E;}
        """)

        excelLayout.addWidget(excelTitulo)
        excelLayout.addWidget(excelDesc)
        excelLayout.addStretch()
        excelLayout.addWidget(self.btnExcel)

        excelCard.setLayout(excelLayout)

        cards.addWidget(pdfCard)
        cards.addWidget(materiasCard)
        cards.addWidget(excelCard)

        layout.addLayout(cards)

        # ==========================
        # INFO EXTRA
        # ==========================
        self.info = QLabel(
            f" Los reportes se guardarán en: {obtener_escritorio()}"
        )

        self.info.setStyleSheet("color:#60A5FA;font-size:13px;padding:10px;")

        layout.addWidget(self.info)
        layout.addStretch()

        self.setLayout(layout)

        self.btnPDF.clicked.connect(self.generar_pdf)
        self.btnPDFMaterias.clicked.connect(self.generar_pdf_materias)
        self.btnExcel.clicked.connect(self.generar_excel)

    # ==================================
    # PDF ALUMNOS
    # ==================================
    def generar_pdf(self):

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from datetime import datetime

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM alumnos")
        alumnos = cursor.fetchall()

        conn.close()

        # GUARDAR EN ESCRITORIO
        escritorio = obtener_escritorio()
        ruta_pdf = os.path.join(escritorio, "Reporte_Alumnos.pdf")

        documento = SimpleDocTemplate(ruta_pdf, pagesize=letter)

        estilos = getSampleStyleSheet()
        elementos = []

        # ENCABEZADO
        titulo = Paragraph(
            """
            <para align='center'>
            <font size='22'><b>SISTEMA DE CONTROL ESCOLAR</b></font>
            <br/><br/>
            <font size='16'>REPORTE GENERAL DE ALUMNOS</font>
            </para>
            """,
            estilos["Title"]
        )

        elementos.append(titulo)
        elementos.append(Spacer(1, 20))

        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

        info = Paragraph(
            f"<b>Fecha de generación:</b> {fecha}<br/><b>Total de alumnos:</b> {len(alumnos)}",
            estilos["Normal"]
        )

        elementos.append(info)
        elementos.append(Spacer(1, 20))

        # TABLA
        datos = [["ID", "Nombre", "Apellido", "Grado", "Grupo"]]

        for alumno in alumnos:
            datos.append([str(alumno[0]), str(alumno[1]), str(alumno[2]), str(alumno[3]), str(alumno[4])])

        tabla = Table(datos, colWidths=[50, 120, 120, 80, 80])

        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
        ]))

        elementos.append(tabla)
        elementos.append(Spacer(1, 30))

        # RESUMEN
        resumen = Paragraph(
            f"<b>Resumen:</b><br/>El presente reporte contiene un total de <b>{len(alumnos)}</b> alumnos registrados en el sistema escolar.",
            estilos["BodyText"]
        )

        elementos.append(resumen)
        elementos.append(Spacer(1, 50))

        # FIRMAS
        firma = Paragraph(
            "<para align='center'>___________________________________<br/>Dirección Escolar</para>",
            estilos["Normal"]
        )

        elementos.append(firma)

        documento.build(elementos)

        QMessageBox.information(
            self,
            "Reporte PDF",
            f"✅ Reporte_Alumnos.pdf guardado en:\n{escritorio}"
        )

    # ==================================
    # PDF MATERIAS
    # ==================================
    def generar_pdf_materias(self):

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("SELECT id, nombre FROM materias ORDER BY nombre")
        materias = cursor.fetchall()

        conn.close()

        if not materias:
            QMessageBox.warning(
                self,
                "Sin materias",
                "No hay materias registradas."
            )
            return

        dialogo = QDialog(self)
        dialogo.setWindowTitle("Seleccionar Materia")
        dialogo.setStyleSheet("""
            QDialog { background-color: #0F172A; color: white; }
            QLabel { color: white; font-size: 16px; }
            QPushButton {
                background-color: #7C3AED;
                border: none;
                border-radius: 10px;
                padding: 12px;
                color: white;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #8B5CF6; }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(20)

        titulo = QLabel(" Selecciona una materia:")
        titulo.setStyleSheet("font-size: 18px; font-weight: bold;")

        layout.addWidget(titulo)

        for m in materias:
            btn = QPushButton(f" {m[1]}")
            btn.clicked.connect(lambda checked, materia_id=m[0], materia_nombre=m[1]: self.generar_pdf_materia_especifica(materia_id, materia_nombre, dialogo))
            layout.addWidget(btn)

        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setStyleSheet("background-color: #EF4444;")
        btn_cancelar.clicked.connect(dialogo.close)
        layout.addWidget(btn_cancelar)

        dialogo.setLayout(layout)
        dialogo.exec()

    # ==================================
    # PDF MATERIA ESPECÍFICA
    # ==================================
    def generar_pdf_materia_especifica(self, materia_id, materia_nombre, dialogo=None):

        if dialogo:
            dialogo.close()

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from datetime import datetime

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT a.id, a.nombre, a.apellido, a.grado, a.grupo
            FROM alumno_materia am
            JOIN alumnos a ON a.id = am.alumno_id
            WHERE am.materia_id = ?
            ORDER BY a.nombre
        """, (materia_id,))

        alumnos = cursor.fetchall()

        conn.close()

        # GUARDAR EN ESCRITORIO
        escritorio = obtener_escritorio()
        ruta_pdf = os.path.join(escritorio, f"Reporte_{materia_nombre}.pdf")

        documento = SimpleDocTemplate(ruta_pdf, pagesize=letter)

        estilos = getSampleStyleSheet()
        elementos = []

        titulo = Paragraph(
            f"""
            <para align='center'>
            <font size='22'><b>SISTEMA DE CONTROL ESCOLAR</b></font>
            <br/><br/>
            <font size='16'>REPORTE: {materia_nombre}</font>
            </para>
            """,
            estilos["Title"]
        )

        elementos.append(titulo)
        elementos.append(Spacer(1, 20))

        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

        info = Paragraph(
            f"<b>Fecha de generación:</b> {fecha}<br/><b>Total de alumnos:</b> {len(alumnos)}",
            estilos["Normal"]
        )

        elementos.append(info)
        elementos.append(Spacer(1, 20))

        if alumnos:
            datos = [["#", "Nombre", "Apellido", "Grado", "Grupo"]]

            for i, alum in enumerate(alumnos, 1):
                datos.append([str(i), str(alum[1]), str(alum[2]), str(alum[3]), str(alum[4])])

            tabla = Table(datos, colWidths=[30, 120, 120, 80, 80])

            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#7C3AED")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
            ]))

            elementos.append(tabla)
        else:
            sin_alumnos = Paragraph(
                "<b>No hay alumnos inscritos en esta materia.</b>",
                estilos["BodyText"]
            )
            elementos.append(sin_alumnos)

        elementos.append(Spacer(1, 30))

        resumen = Paragraph(
            f"<b>Resumen:</b><br/>La materia <b>{materia_nombre}</b> tiene <b>{len(alumnos)}</b> alumnos inscritos.",
            estilos["BodyText"]
        )

        elementos.append(resumen)
        elementos.append(Spacer(1, 50))

        firma = Paragraph(
            "<para align='center'>___________________________________<br/>Dirección Escolar</para>",
            estilos["Normal"]
        )

        elementos.append(firma)

        documento.build(elementos)

        QMessageBox.information(
            self,
            "Reporte Generado",
            f"✅ Reporte_{materia_nombre}.pdf guardado en:\n{escritorio}"
        )

    # ==================================
    # EXCEL
    # ==================================
        # ==================================
    # EXCEL
    # ==================================
        # ==================================
    # EXCEL
    # ==================================
    def generar_excel(self):

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("SELECT id, nombre, apellido, grado, grupo FROM alumnos")
        alumnos = cursor.fetchall()

        conn.close()

        # GUARDAR EN ESCRITORIO
        escritorio = obtener_escritorio()
        ruta_excel = os.path.join(escritorio, "Alumnos.xlsx")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Alumnos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_align = Alignment(horizontal="center")

        # Encabezados
        headers = ["ID", "Nombre", "Apellido", "Grado", "Grupo"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Datos
        for row, alum in enumerate(alumnos, 2):
            ws.cell(row=row, column=1, value=alum[0] if alum[0] else "")
            ws.cell(row=row, column=2, value=alum[1] if alum[1] else "")
            ws.cell(row=row, column=3, value=alum[2] if alum[2] else "")
            ws.cell(row=row, column=4, value=alum[3] if alum[3] else "")
            ws.cell(row=row, column=5, value=alum[4] if alum[4] else "")

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        # Guardar
        wb.save(ruta_excel)

        QMessageBox.information(
            self,
            "Reporte Excel",
            f"✅ Alumnos.xlsx guardado en:\n{escritorio}"
        )